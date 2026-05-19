from __future__ import annotations

import argparse
import csv
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CONTRIBUTION_LABELS = [
    "artifact_dataset",
    "artifact_method",
    "artifact_task",
    "knowledge_dataset",
    "knowledge_language",
    "knowledge_method",
    "knowledge_people",
    "knowledge_task",
]

VENUE_LABELS = {
    "acl": "ACL",
    "emnlp": "EMNLP",
    "naacl": "NAACL",
}


DEFAULT_CONTRIBUTIONS = Path(
    r"C:\Users\shuaichen\Desktop\Chen Shuai\code\computing_resource"
    r"\data\processed\nlp_contributions\acl_scibert_fixed_core_contributions_t08_papers.csv"
)

DEFAULT_COMPUTE_FILES = [
    Path(
        r"C:\Users\shuaichen\Desktop\Chen Shuai\code\computing_resource"
        r"\analysis\data\data\soft_compute_paper_level.xlsx"
    ),
    Path(
        r"C:\Users\shuaichen\Desktop\Chen Shuai\code\computing_resource"
        r"\analysis\data\data\strict_compute_paper_level.xlsx"
    ),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append NLP contribution fields to paper-level compute workbooks."
    )
    parser.add_argument("--contributions", type=Path, default=DEFAULT_CONTRIBUTIONS)
    parser.add_argument("--compute-file", type=Path, action="append", dest="compute_files")
    parser.add_argument("--output-suffix", default="_with_contributions")
    return parser.parse_args()


def load_contributions(path: Path) -> dict[str, dict[str, str]]:
    records: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            anthology_id = row.get("anthology_id", "").strip()
            if anthology_id:
                records[anthology_id] = row
    return records


def contribution_columns() -> list[str]:
    columns = [
        "contribution_matched",
        "contribution_analysis_included",
        "paper_year",
        "paper_venue",
        "contribution_title",
        "contribution_year",
        "contribution_venue",
        "contribution_url",
        "contribution_pdf_url",
        "contribution_threshold",
        "contribution_selection_status",
        "contribution_num_sentences",
        "contribution_labels_at_threshold",
        "contribution_core_labels",
        "contribution_core_sentence_count",
        "contribution_core_contributions",
    ]
    columns.extend(f"contribution_core_{label}" for label in CONTRIBUTION_LABELS)
    return columns


def parse_paper_id_year_venue(paper_id: str) -> tuple[str, str]:
    parts = paper_id.strip().split(".")
    if len(parts) < 3:
        return "", ""

    year = parts[0] if parts[0].isdigit() and len(parts[0]) == 4 else ""
    venue_part = ".".join(parts[1:-1])
    venue_code = venue_part.removeprefix("findings-").split("-")[0]
    venue = VENUE_LABELS.get(venue_code.lower(), venue_code.upper() if venue_code else "")
    return year, venue


def contribution_values(record: dict[str, str] | None, paper_id: str) -> list[Any]:
    paper_year, paper_venue = parse_paper_id_year_venue(paper_id)
    if record is None:
        return [
            0,
            0,
            paper_year,
            paper_venue,
            "",
            paper_year,
            paper_venue,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            None,
            "",
        ] + [None] * len(CONTRIBUTION_LABELS)

    core_labels = {
        label.strip()
        for label in record.get("core_labels", "").split(";")
        if label.strip()
    }
    return [
        1,
        1,
        paper_year,
        paper_venue,
        record.get("title", ""),
        record.get("year", "") or paper_year,
        record.get("venue", "") or paper_venue,
        record.get("url", ""),
        record.get("pdf_url", ""),
        record.get("threshold", ""),
        record.get("selection_status", ""),
        record.get("num_sentences", ""),
        record.get("labels_at_threshold", ""),
        record.get("core_labels", ""),
        int(record.get("core_sentence_count", "0") or 0),
        record.get("core_contributions", ""),
        *[1 if label in core_labels else 0 for label in CONTRIBUTION_LABELS],
    ]


def find_header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def style_new_header(sheet, col_idx: int) -> None:
    cell = sheet.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def copy_adjacent_style(sheet, row_idx: int, col_idx: int) -> None:
    source = sheet.cell(row=row_idx, column=col_idx - 1)
    target = sheet.cell(row=row_idx, column=col_idx)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def merge_workbook(
    compute_file: Path,
    contributions: dict[str, dict[str, str]],
    output_suffix: str,
) -> tuple[Path, int, int]:
    workbook = load_workbook(compute_file)
    sheet = workbook.active
    header_map = find_header_map(sheet)
    if "paper_id" not in header_map:
        raise ValueError(f"{compute_file} does not contain a paper_id column")

    start_col = sheet.max_column + 1
    columns = contribution_columns()
    for offset, column_name in enumerate(columns):
        col_idx = start_col + offset
        sheet.cell(row=1, column=col_idx, value=column_name)
        style_new_header(sheet, col_idx)

    paper_id_col = header_map["paper_id"]
    matched = 0
    for row_idx in range(2, sheet.max_row + 1):
        paper_id = str(sheet.cell(row=row_idx, column=paper_id_col).value or "").strip()
        record = contributions.get(paper_id)
        if record is not None:
            matched += 1
        for offset, value in enumerate(contribution_values(record, paper_id)):
            col_idx = start_col + offset
            copy_adjacent_style(sheet, row_idx, col_idx)
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if columns[offset] == "contribution_core_contributions":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "contribution_matched": 16,
        "contribution_analysis_included": 28,
        "paper_year": 12,
        "paper_venue": 14,
        "contribution_title": 48,
        "contribution_year": 12,
        "contribution_venue": 14,
        "contribution_url": 42,
        "contribution_pdf_url": 42,
        "contribution_threshold": 16,
        "contribution_selection_status": 22,
        "contribution_num_sentences": 18,
        "contribution_labels_at_threshold": 42,
        "contribution_core_labels": 42,
        "contribution_core_sentence_count": 24,
        "contribution_core_contributions": 100,
    }
    widths.update({f"contribution_core_{label}": 18 for label in CONTRIBUTION_LABELS})
    for offset, column_name in enumerate(columns):
        col_letter = sheet.cell(row=1, column=start_col + offset).column_letter
        sheet.column_dimensions[col_letter].width = widths.get(column_name, 18)

    if sheet.max_row > 1 and sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "B2"

    output_file = compute_file.with_name(f"{compute_file.stem}{output_suffix}{compute_file.suffix}")
    workbook.save(output_file)
    return output_file, matched, sheet.max_row - 1


def main() -> None:
    args = parse_args()
    compute_files = args.compute_files or DEFAULT_COMPUTE_FILES
    contributions = load_contributions(args.contributions)
    print(f"loaded_contributions={len(contributions)}")
    for compute_file in compute_files:
        output_file, matched, rows = merge_workbook(
            compute_file=compute_file,
            contributions=contributions,
            output_suffix=args.output_suffix,
        )
        print(f"{compute_file.name}: rows={rows} matched={matched} unmatched={rows - matched}")
        print(f"wrote={output_file}")


if __name__ == "__main__":
    main()
