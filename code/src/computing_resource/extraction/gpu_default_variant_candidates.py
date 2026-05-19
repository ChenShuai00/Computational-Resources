from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from computing_resource.extraction.gpu_catalog import load_hardware_catalog, resolve_hardware_name
from computing_resource.extraction.gpu_excel_export import EXPORT_HEADERS


ROW_HEADERS = [
    "source_file",
    "gpu_name",
    "gpu_num",
    "current_benchmark_gpu_name",
    "normalize_status",
    "normalize_reason",
    "catalog_candidate_count",
    "catalog_candidates",
    "candidate_reason",
]

UNIQUE_HEADERS = [
    "gpu_name",
    "row_count",
    "file_count",
    "catalog_candidate_count",
    "catalog_candidates",
    "candidate_reason",
    "sample_source_files",
    "default_benchmark_gpu_name",
]


def _header_index() -> dict[str, int]:
    return {name: index for index, name in enumerate(EXPORT_HEADERS)}


def _candidate_reason(row: tuple[Any, ...], resolution) -> str | None:
    index = _header_index()
    if row[index["normalize_status"]] == "normalized" and not row[index["benchmark_gpu_name"]]:
        return "normalized_without_benchmark"
    if len(resolution.candidate_names) > 1:
        return "multiple_catalog_candidates"
    return None


def export_default_variant_candidates(
    input_path: str | Path,
    catalog_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    source = Path(input_path)
    target = Path(output_path) if output_path is not None else source.with_name(
        f"{source.stem}_default_variant_candidates{source.suffix}"
    )

    workbook = load_workbook(source, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if headers != EXPORT_HEADERS:
        raise ValueError("Input workbook headers do not match normalized GPU export schema")

    catalog = load_hardware_catalog(catalog_path)
    index = _header_index()
    row_level: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        gpu_name = row[index["gpu_name"]]
        if not gpu_name:
            continue
        resolution = resolve_hardware_name(gpu_name, catalog)
        reason = _candidate_reason(row, resolution)
        if reason is None:
            continue
        row_level.append(
            {
                "source_file": row[index["source_file"]],
                "gpu_name": gpu_name,
                "gpu_num": row[index["gpu_num"]],
                "current_benchmark_gpu_name": row[index["benchmark_gpu_name"]],
                "normalize_status": row[index["normalize_status"]],
                "normalize_reason": row[index["normalize_reason"]],
                "catalog_candidate_count": len(resolution.candidate_names),
                "catalog_candidates": "; ".join(resolution.candidate_names),
                "candidate_reason": reason,
            }
        )

    unique_gpu_names: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "row_count": 0,
            "source_files": set(),
            "catalog_candidate_count": 0,
            "catalog_candidates": set(),
            "candidate_reason": set(),
        }
    )
    for row in row_level:
        bucket = unique_gpu_names[row["gpu_name"]]
        bucket["row_count"] += 1
        bucket["source_files"].add(row["source_file"])
        bucket["catalog_candidate_count"] = max(
            bucket["catalog_candidate_count"], row["catalog_candidate_count"]
        )
        if row["catalog_candidates"]:
            bucket["catalog_candidates"].add(row["catalog_candidates"])
        bucket["candidate_reason"].add(row["candidate_reason"])

    output = Workbook()
    row_sheet = output.active
    row_sheet.title = "row_level"
    row_sheet.append(ROW_HEADERS)
    for row in row_level:
        row_sheet.append([row.get(header) for header in ROW_HEADERS])

    unique_sheet = output.create_sheet("unique_gpu_names")
    unique_sheet.append(UNIQUE_HEADERS)
    for gpu_name, row in sorted(
        unique_gpu_names.items(),
        key=lambda item: (-item[1]["row_count"], item[0].lower()),
    ):
        unique_sheet.append(
            [
                gpu_name,
                row["row_count"],
                len(row["source_files"]),
                row["catalog_candidate_count"],
                " | ".join(sorted(row["catalog_candidates"])),
                " | ".join(sorted(row["candidate_reason"])),
                " | ".join(sorted(row["source_files"])[:5]),
                "",
            ]
        )

    target.parent.mkdir(parents=True, exist_ok=True)
    output.save(target)
    return target
