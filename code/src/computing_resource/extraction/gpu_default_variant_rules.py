from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import yaml


def load_default_variant_rules(rules_path: str | Path | None) -> dict[str, str]:
    if rules_path is None:
        return {}
    path = Path(rules_path)
    if not path.exists():
        raise FileNotFoundError(f"Default variant rules file does not exist: {path}")
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    mapping = data.get("default_benchmark_variants", {}) or {}
    return {str(key): str(value) for key, value in mapping.items() if value}


def export_default_variant_rules(input_path: str | Path, output_path: str | Path) -> Path:
    source = Path(input_path)
    target = Path(output_path)

    workbook = load_workbook(source, read_only=True)
    worksheet = workbook["unique_gpu_names"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(headers)}

    rules: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        gpu_name = row[index["gpu_name"]]
        default_name = row[index["default_benchmark_gpu_name"]]
        if gpu_name and default_name:
            rules[str(gpu_name)] = str(default_name)

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        yaml.safe_dump(
            {"default_benchmark_variants": rules},
            allow_unicode=True,
            sort_keys=True,
        ),
        encoding="utf-8",
    )
    return target
