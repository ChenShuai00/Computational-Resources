from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from computing_resource.extraction.gpu_catalog import (
    MODEL_OR_FRAMEWORK_NAME_REASON,
    load_hardware_catalog,
    resolve_hardware_name,
)
from computing_resource.extraction.gpu_default_variant_rules import load_default_variant_rules
from computing_resource.extraction.gpu_excel_export import EXPORT_HEADERS, _benchmark_columns_for_row


HEADER_INDEX = {name: index for index, name in enumerate(EXPORT_HEADERS)}
DEFAULT_VARIANT_RULES_PATH = Path(__file__).resolve().parents[3] / "config" / "gpu_default_benchmark_variants.yaml"
CPU_FALLBACK_BENCHMARK_NAME = "cpu"
CPU_FALLBACK_REASON = "cpu_fallback"


def _normalize_status(match_status: str | None) -> str:
    if match_status in {"unmatched", "ambiguous"}:
        return "unresolved"
    return "normalized"


def _default_variant_benchmark_name(resolution, default_variant_rules: dict[str, str]) -> str | None:
    if resolution.benchmark_hardware_name:
        return None

    candidate_keys = (
        resolution.raw_hardware_name,
        resolution.normalized_hardware_name,
        resolution.cleaned_hardware_name,
    )
    for candidate_key in candidate_keys:
        if candidate_key and candidate_key in default_variant_rules:
            return default_variant_rules[candidate_key]
    return None


def _cpu_fallback_benchmark_name(resolution) -> str | None:
    if resolution.benchmark_hardware_name:
        return None

    normalized_text = resolution.cleaned_hardware_name or ""
    tokens = set(normalized_text.split())
    if not normalized_text:
        return None

    gpu_exclusion_tokens = {"arc", "cuda", "geforce", "gpu", "gtx", "instinct", "iris", "quadro", "radeon", "rtx", "tesla", "tpu"}
    if tokens.intersection(gpu_exclusion_tokens):
        return None
    if normalized_text in {"amd", "intel", "cpu", "cpus"}:
        return CPU_FALLBACK_BENCHMARK_NAME
    if any(
        re.search(pattern, normalized_text)
        for pattern in (
            r"\bcpu\b",
            r"\bxeon\b",
            r"\bepyc\b",
            r"\bryzen\b",
            r"\bthreadripper\b",
            r"\bintel\s+core\b",
            r"\bintel\s+i[3579]\b",
        )
    ):
        return CPU_FALLBACK_BENCHMARK_NAME
    if ("amd" in tokens or "intel" in tokens) and not tokens.intersection(gpu_exclusion_tokens):
        return CPU_FALLBACK_BENCHMARK_NAME
    return None


def _updated_row_values(values: list[Any], catalog, default_variant_rules: dict[str, str]) -> list[Any]:
    row = list(values)
    raw_hardware_name = row[HEADER_INDEX["gpu_name"]]
    resolution = resolve_hardware_name(raw_hardware_name or "", catalog)
    default_variant_benchmark_name = _default_variant_benchmark_name(resolution, default_variant_rules)
    cpu_fallback_benchmark_name = _cpu_fallback_benchmark_name(resolution)
    benchmark_name = default_variant_benchmark_name or resolution.benchmark_hardware_name or cpu_fallback_benchmark_name
    benchmark_row = catalog.row_map.get(benchmark_name) if benchmark_name else None
    if default_variant_benchmark_name and benchmark_row is None:
        raise ValueError(
            f"Default variant rule resolved to an unknown benchmark: {default_variant_benchmark_name}"
        )
    if cpu_fallback_benchmark_name and benchmark_row is None:
        benchmark_row = {"Hardware name": CPU_FALLBACK_BENCHMARK_NAME}
    benchmark = _benchmark_columns_for_row(benchmark_row)

    row[HEADER_INDEX["benchmark_gpu_name"]] = benchmark.name
    row[HEADER_INDEX["gpu_vendor"]] = benchmark.vendor
    row[HEADER_INDEX["benchmark_generation"]] = benchmark.generation
    row[HEADER_INDEX["benchmark_family"]] = benchmark.family
    row[HEADER_INDEX["benchmark_release_date"]] = benchmark.release_date
    row[HEADER_INDEX["benchmark_release_price_usd"]] = benchmark.release_price_usd
    row[HEADER_INDEX["benchmark_tdp_w"]] = benchmark.tdp_w
    row[HEADER_INDEX["benchmark_memory_bytes"]] = benchmark.memory_bytes
    row[HEADER_INDEX["benchmark_memory_bandwidth_bytes_per_s"]] = benchmark.memory_bandwidth_bytes_per_s
    row[HEADER_INDEX["benchmark_fp32_flops"]] = benchmark.fp32_flops
    row[HEADER_INDEX["benchmark_tf32_flops"]] = benchmark.tf32_flops
    row[HEADER_INDEX["benchmark_tensor_fp16_bf16_flops"]] = benchmark.tensor_fp16_bf16_flops
    row[HEADER_INDEX["benchmark_fp8_flops"]] = benchmark.fp8_flops
    row[HEADER_INDEX["benchmark_max_performance"]] = benchmark.max_performance
    row[HEADER_INDEX["benchmark_energy_efficiency"]] = benchmark.energy_efficiency
    row[HEADER_INDEX["benchmark_process_size_nm"]] = benchmark.process_size_nm
    row[HEADER_INDEX["normalize_status"]] = (
        "normalized" if benchmark_name else _normalize_status(resolution.match_status)
    )
    row[HEADER_INDEX["normalize_reason"]] = (
        "manual_default_variant_rule"
        if default_variant_benchmark_name
        else CPU_FALLBACK_REASON
        if cpu_fallback_benchmark_name
        else resolution.normalization_reason
    )
    return row


def _should_remove_row(values: list[Any]) -> bool:
    return (
        values[HEADER_INDEX["normalize_reason"]] == MODEL_OR_FRAMEWORK_NAME_REASON
        or values[HEADER_INDEX["benchmark_gpu_name"]] == CPU_FALLBACK_BENCHMARK_NAME
    )


def renormalize_gpu_excel(
    input_path: str | Path,
    catalog_path: str | Path,
    output_path: str | Path | None = None,
    in_place: bool = False,
    default_variant_rules_path: str | Path | None = None,
) -> Path:
    source = Path(input_path)
    if in_place and output_path is not None:
        raise ValueError("Use either in_place or output_path, not both")
    target = source if in_place else Path(output_path) if output_path is not None else source.with_name(
        f"{source.stem}_normalized{source.suffix}"
    )

    catalog = load_hardware_catalog(catalog_path)
    rules_path = default_variant_rules_path or (
        DEFAULT_VARIANT_RULES_PATH if DEFAULT_VARIANT_RULES_PATH.exists() else None
    )
    default_variant_rules = load_default_variant_rules(rules_path)
    workbook = load_workbook(source)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if headers != EXPORT_HEADERS:
        raise ValueError("Input workbook headers do not match normalized GPU export schema")

    rows_to_delete: list[int] = []
    for row_index, row_values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        updated_values = _updated_row_values(list(row_values), catalog, default_variant_rules)
        if _should_remove_row(updated_values):
            rows_to_delete.append(row_index)
            continue
        for column_index, value in enumerate(updated_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    for row_index in reversed(rows_to_delete):
        worksheet.delete_rows(row_index, 1)

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
